from openpyxl import load_workbook
import csv 

workbook = load_workbook(filename="./files/tickers.xlsx")
sheet = workbook.active

tickers_list = []

for row in sheet.iter_rows(min_row=5, max_col=5, values_only=True):
    if row[4] == 'USA':
        tickers_list.append(row[0])

# name of csv file 
filename = "./files/unique_usa_tickers.csv"
    
# writing to csv file 
with open(filename, 'w', newline='') as csvfile: 
    # creating a csv writer object 
    csvwriter = csv.writer(csvfile)     
    # writing the data rows 
    for ticker in tickers_list:
        csvwriter.writerow([ticker])

print("Finished Successfully!")